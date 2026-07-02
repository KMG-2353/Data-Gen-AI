"""End-to-end defect coverage for the 2026-06-30 Defect Tracker, run against the
re-structured (numbered-sheet) SPG blank templates.

Each test reproduces a logged defect by feeding the handler the exact bad row the
LLM would emit (or, for the universal phone/state passes, the row main.py hands to
``enforce_variety_fields``) and asserts the deterministic engine repairs it.

Defect map:
  IM    : DF-IM-013 (entity), DF-IM-017 (equipment multi), DF-IM-018 (loss payees)
  Cargo : CARGO-005 (entity)
  APD   : APD-004 (entity), APD-005 (company), APD-006/007/008/009 (units multi),
          APD-010 (loss payees), APD-011 (loss history), APD-012 (scenario sheet)
  WH    : WH-001 (phone), WH-002 (state, incl the "St" column)
  DW    : DEF-016 (DF Locations >2 Acres, DF Loss Payees State)
  HO    : HO-018 (>2 Acres), HO-019 (Number of Golf Carts), HO-020 (Is Mortgagee)
"""
import os
import random
from collections import Counter

import openpyxl
import pytest

from app.llm_service import detect_policy_type_from_headers
from app.policies import get_handler
from app.policies.im import ImHandler
from app.policies.spg_auto import ApdHandler, CargoHandler, _AUTO_ENTITY_TYPES, _AUTO_COMPANY_TYPES
from app.policies.spg_wh import WhHandler, _WH_ENTITY_TYPES
from app.policies.spg_pl import DwHandler, HoHandler, _PL_ENTITY_TYPES
from app.policies.im import _IM_ENTITY_TYPES
from app.rulebook.variety import enforce_variety_fields
from app.rulebook.primitives import is_us_state, parse_date as _parse_date


@pytest.fixture(autouse=True)
def _seed():
    random.seed(20260630)


_BLANK_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "Rater Files", "Blank Templates"
)


def _headers_by_sheet(filename: str) -> dict:
    wb = openpyxl.load_workbook(os.path.join(_BLANK_DIR, filename), read_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if len(cells) >= 2:
                out[sn] = cells
                break
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Detection: every re-structured template routes to its handler (not GENERIC)
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("filename,expected", [
    ("SPG_IM_Sample_Headers.xlsx", "IM"),
    ("SPG_Cargo_Sample_Headers.xlsx", "CARGO"),
    ("SPG_APD_Sample_Headers.xlsx", "APD"),
    ("SPG_WH_Sample_Headers.xlsx", "WH"),
    ("PersonalLine_DW_Rater_v0.01.xlsx", "DW"),
    ("PersonalLine_HO_Rater_v0.01.xlsx", "HO"),
])
def test_new_templates_detect(filename, expected):
    assert detect_policy_type_from_headers(_headers_by_sheet(filename)) == expected


def test_wh_handler_registered():
    assert get_handler("WH").policy_type == "WH"


# ---------------------------------------------------------------------------
# IM (new template) — DF-IM-013 / 017 / 018
# ---------------------------------------------------------------------------

def test_im_entity_varies_new_template():
    # DF-IM-013: the LLM collapses "Type of Entity" onto Corporation/LLC.
    handler = ImHandler()
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Effective Date": "05/27/2026", "Quote Date": "05/20/2026"} for _ in range(6)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    seen = {r["Type of Entity"] for r in out}
    assert len(seen) >= 4
    assert seen <= set(_IM_ENTITY_TYPES)


def test_im_equipment_multi_row():
    # DF-IM-017: a single equipment row must be expanded to several (<= 20).
    handler = ImHandler()
    rows = [{"Test ID": "TS-001", "#": 1, "Serial Number": "SN1", "Value ($)": "$30,000"}]
    out = handler.post_process(rows, "03_IM_Equipment", "",
                               {"01_Policy_Info": [{"Test ID": "TS-001"}]})
    assert len(out) >= 4
    assert len({r["Serial Number"] for r in out}) == len(out)   # uniquified
    assert all(25000 < int(r["Value ($)"]) <= 50000 for r in out)


def test_im_loss_payees_varied_yes_no_and_blank():
    # DF-IM-018: "Has Loss Payees?" must VARY Yes/No across insureds — the LLM
    # collapses it onto Yes for everyone. Yes insureds carry a 1–10 multi-row
    # schedule; No insureds carry a SINGLE row with every detail (and #) blank.
    handler = ImHandler()
    roster = [{"Test ID": f"TS-{i:03d}"} for i in range(1, 7)]
    rows = [{"Test ID": f"TS-{i:03d}", "Has Loss Payees?  (Yes / No)": "Yes", "#": 1,
             "Loss Payee Name": "Bank", "State": "VA", "Loan Number": f"L{i}"}
            for i in range(1, 7)]
    out = handler.post_process(rows, "05_IM_LossPayees", "", {"01_Policy_Info": roster})

    flag_key = "Has Loss Payees?  (Yes / No)"
    flags = {r["Test ID"]: r[flag_key] for r in out}
    assert set(flags.values()) == {"Yes", "No"}            # both values appear

    counts = Counter(r["Test ID"] for r in out)
    for tid, f in flags.items():
        if f == "No":
            assert counts[tid] == 1                        # single row
            row = next(r for r in out if r["Test ID"] == tid)
            assert row["Loss Payee Name"] == ""            # blank detail …
            assert row["Loan Number"] == ""
            assert row["#"] == ""                          # … and blank #
        else:
            assert counts[tid] >= 2                        # multi-row schedule


def test_im_misc_total_blank_when_disabled():
    handler = ImHandler()
    out = handler.post_process(
        [{"Test ID": "x", "Enable Miscellaneous Articles?": "No",
          "Total Value of Miscellaneous Articles ($)": "5000"}],
        "04_IM_MiscArticles", "")
    assert out[0]["Total Value of Miscellaneous Articles ($)"] == ""


# ---------------------------------------------------------------------------
# Cargo / APD — entity, company, child multiplicity, scenario sheet
# ---------------------------------------------------------------------------

def test_cargo_entity_varies():
    # CARGO-005
    handler = CargoHandler()
    rows = [{"Test ID": "x", "Type of Entity": "LLC"} for _ in range(7)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    seen = {r["Type of Entity"] for r in out}
    assert len(seen) >= 5 and seen <= set(_AUTO_ENTITY_TYPES)


def test_apd_entity_and_company():
    # APD-004 (entity) + APD-005 (Type of Company / Carrier "tow trucks" invalid)
    handler = ApdHandler()
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Type of Company / Carrier": "tow trucks"} for _ in range(7)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    assert {r["Type of Entity"] for r in out} <= set(_AUTO_ENTITY_TYPES)
    assert len({r["Type of Entity"] for r in out}) >= 5
    for r in out:
        assert r["Type of Company / Carrier"] in _AUTO_COMPANY_TYPES   # "tow trucks" snapped out


@pytest.mark.parametrize("sheet,colname,unique", [
    ("03_APD_Drivers", "License Number", True),     # APD-006
    ("04_APD_Vehicles", "VIN Number", True),        # APD-007
    ("05_APD_Trailers", "VIN Number", True),        # APD-008
    ("06_APD_Commodities", "Commodity (Select from list)", False),  # APD-009
])
def test_apd_units_multi_row(sheet, colname, unique):
    handler = ApdHandler()
    rows = [{"Test ID": "TS-001", colname: "A"}]
    out = handler.post_process(rows, sheet, "", {"01_Policy_Info": [{"Test ID": "TS-001"}]})
    assert len(out) >= 2                              # expanded beyond the single LLM row
    assert all(r["Test ID"] == "TS-001" for r in out)


def test_apd_loss_payees_and_history():
    handler = ApdHandler()
    prev = {"01_Policy_Info": [{"Test ID": "TS-001"}, {"Test ID": "TS-002"}]}
    # APD-010
    lp = [{"Test ID": "TS-001", "Does Policy Have Loss Payees (PhysDam)?": "Yes",
           "Loss Payee Name": "Bank", "State": "VA"}]
    olp = handler.post_process(lp, "09_APD_LossPayees", "", prev)
    assert len(olp) >= 2
    # APD-011: Yes expands, No stays single + blanked
    lh = [{"Test ID": "TS-001", "Any Losses in the Past 3 Years?": "Yes", "Loss Year": "2024"},
          {"Test ID": "TS-002", "Any Losses in the Past 3 Years?": "No", "Loss Year": "2023"}]
    olh = handler.post_process(lh, "08_APD_LossHistory", "", prev)
    counts = Counter(r["Test ID"] for r in olh)
    assert counts["TS-001"] >= 2 and counts["TS-002"] == 1
    assert next(r for r in olh if r["Test ID"] == "TS-002")["Loss Year"] == ""


def test_apd_scenario_details_built():
    # APD-012: the Test Scenario Details summary is built deterministically.
    handler = ApdHandler()
    prev = {
        "01_Policy_Info": [{"Test ID": "TS-001", "Binding State": "VA", "Type of Entity": "LLC"}],
        "03_APD_Drivers": [{"Test ID": "TS-001"}] * 3,
        "04_APD_Vehicles": [{"Test ID": "TS-001"}] * 2,
        "09_APD_LossPayees": [{"Test ID": "TS-001"}] * 2,
    }
    headers = ["Scenario ID", "State", "Type of Entity", "Driver Count",
               "Vehicle Count", "Trailer Count", "Commodity Count",
               "Loss History Count", "Loss Payee Count"]
    out = handler.pre_generate("Test Scenario Details", headers,
                               prev["01_Policy_Info"], None, None, prev)
    assert out and out[0]["Scenario ID"] == "TS-001"
    assert out[0]["Driver Count"] == 3
    assert out[0]["Vehicle Count"] == 2
    assert out[0]["Loss Payee Count"] == 2


# ---------------------------------------------------------------------------
# Wind / Hail — WH-001 (phone), WH-002 (state incl "St"), entity, exp
# ---------------------------------------------------------------------------

def test_wh_policy_entity_exp_quote():
    from datetime import date
    from app.rulebook.primitives import format_date_slash
    handler = WhHandler()
    # Future effective date so it is preserved (not clamped up to today).
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Effective Date": "12/01/2027", "Expiration Date": "",
             "Quote Date": "06/10/2026"} for _ in range(6)]
    out = handler.post_process([dict(r) for r in rows], "01_Policy_Info", "")
    assert {r["Type of Entity"] for r in out} <= set(_WH_ENTITY_TYPES)
    assert len({r["Type of Entity"] for r in out}) >= 4
    assert out[0]["Expiration Date"] == "12/01/2028"      # eff + 1 year
    # WH-005: Quote Date pinned to today (data-creation date).
    assert out[0]["Quote Date"] == format_date_slash(date.today())


def test_wh_phone_and_state_variety():
    # WH-001 (phone format) + WH-002 (state spread, including the "St" column).
    rows = [
        {"Test ID": "TS-001", "Binding State  (auto from Location 1)": "VA",
         "St": "VA", "Address - State": "VA", "Agent Phone": "8045551234",
         "Agent Fax": "8045559999"},
        {"Test ID": "TS-001", "Binding State  (auto from Location 1)": "VA",
         "St": "VA", "Address - State": "VA", "Agent Phone": "8045551235",
         "Agent Fax": "8045559998"},
    ]
    enforce_variety_fields(rows)
    for r in rows:
        assert r["Agent Phone"].startswith("(")            # WH-001
        assert r["Agent Fax"].startswith("(")
        assert is_us_state(r["St"])                        # WH-002
    assert {r["St"] for r in rows} != {"VA"}               # not collapsed onto binding


def test_wh_locations_multi_row():
    handler = WhHandler()
    rows = [{"Test ID": "TS-001", "#": 1, "St": "VA", "City": "X"}]
    out = handler.post_process(rows, "02_WH_Locations", "",
                               {"01_Policy_Info": [{"Test ID": "TS-001"}]})
    assert len(out) >= 4
    assert [r["#"] for r in out] == list(range(1, len(out) + 1))


# ---------------------------------------------------------------------------
# DW / HO collapse defects
# ---------------------------------------------------------------------------

def test_dw_two_acres_varies():
    # DEF-016 (DF Locations "> 2 Acres?" always No)
    handler = DwHandler()
    rows = [{"Test ID": "TS-001", "> 2 Acres?": "No", "> 10 Acres?": "No", "State": "VA"}
            for _ in range(6)]
    out = handler.post_process(rows, "DF Locations", "", {"Policy Info": [{"Test ID": "TS-001"}]})
    seen = {r["> 2 Acres?"] for r in out}
    assert "Yes" in seen and "No" in seen
    # dependency: > 10 Acres blanked on the No rows
    for r in out:
        if str(r["> 2 Acres?"]).lower() == "no":
            assert r["> 10 Acres?"] == ""


def test_dw_locations_expanded_to_multiple():
    # DEF-007: a single DF location must be expanded to several (<= 20).
    handler = DwHandler()
    rows = [{"Test ID": "TS-001", "Loc #": 1, "State": "VA", "> 2 Acres?": "No",
             "> 10 Acres?": "No", "Single Family?": "Yes", "Siding Material": "Vinyl"}]
    out = handler.post_process(rows, "DF Locations", "",
                               {"Policy Info": [{"Test ID": "TS-001", "Effective Date": "05/27/2026"}]})
    assert len([r for r in out if r["Test ID"] == "TS-001"]) >= 4


def test_dw_loss_history_multi_when_losses_single_when_none():
    # DEF-010: a loss policy gets a multi-row schedule; a no-loss policy stays one row.
    handler = DwHandler()
    yes = handler.post_process(
        [{"Test ID": "TS-001", "Any Losses in Past 5 Years?": "Yes", "#": 1,
          "Loss Date": "01/15/2025", "Amount": "5000"}],
        "DF Loss history", "",
        {"Policy Info": [{"Test ID": "TS-001", "Effective Date": "05/27/2026"}]})
    assert len(yes) >= 2
    no = handler.post_process(
        [{"Test ID": "TS-002", "Any Losses in Past 5 Years?": "No", "#": 1,
          "Loss Date": "", "Amount": ""}],
        "DF Loss history", "",
        {"Policy Info": [{"Test ID": "TS-002", "Effective Date": "05/27/2026"}]})
    assert len(no) == 1


def test_dw_losspayee_state_varies():
    # DEF-016 (DF Loss Payees "State" only from binding dropdown)
    handler = DwHandler()
    rows = [{"Test ID": "TS-001", "State": "VA", "Binding State": "VA",
             "Is Mortgagee?": "Yes", "Loan Number": "L"} for _ in range(6)]
    out = handler.post_process(rows, "DF Loss Payees", "", {"Policy Info": [{"Test ID": "TS-001"}]})
    enforce_variety_fields(out)
    assert all(is_us_state(r["State"]) for r in out)
    assert {r["State"] for r in out} != {"VA"}


def test_ho_two_acres_and_golf_carts():
    # HO-018 (>2 Acres always No) + HO-019 (Number of Golf Carts always 0)
    handler = HoHandler()
    rows = [{"Test ID": "TS-001", "Located on More Than 2 Acres?": "No",
             "Located on More Than 10 Acres?": "No", "Number of Golf Carts": 0,
             "Protection Class": 5} for _ in range(6)]
    out = handler.post_process(rows, "HO Dwelling", "", {"Policy Info": [{"Test ID": "TS-001"}]})
    assert {r["Located on More Than 2 Acres?"] for r in out} == {"Yes", "No"}
    assert any(int(r["Number of Golf Carts"]) > 0 for r in out)
    assert any(int(r["Number of Golf Carts"]) == 0 for r in out)


def test_ho_mortgagee_varies():
    # HO-020 (Is Mortgagee? always Yes)
    handler = HoHandler()
    rows = [{"Test ID": "TS-001", "Is Mortgagee?": "Yes", "Loan Number": "L1",
             "Mortgage Current?": "Yes", "State": "VA"} for _ in range(6)]
    out = handler.post_process(rows, "HO Loss Payees", "", {"Policy Info": [{"Test ID": "TS-001"}]})
    seen = {r["Is Mortgagee?"] for r in out}
    assert seen == {"Yes", "No"}
    for r in out:
        if str(r["Is Mortgagee?"]).lower() == "no":
            assert r["Loan Number"] == ""               # dependency enforced


# ---------------------------------------------------------------------------
# Class A — Effective Date >= Quote Date (coverage cannot begin before the
# quote is created).  DF-IM-020 / DEF-023 / HO-022 / CARGO-006 / WH-003 / APD-013
# ---------------------------------------------------------------------------
import datetime as _dt


def _eff_ge_quote(row):
    eff = _parse_date(row.get("Effective Date") or row.get("effective date"))
    quote = _parse_date(row.get("Quote Date") or row.get("Date of Quote"))
    assert eff is not None and quote is not None
    assert eff >= quote, f"effective {eff} < quote {quote}"


def test_im_effective_on_or_after_quote():
    # DF-IM-020: LLM emitted an effective date earlier than the quote date.
    handler = ImHandler()
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Effective Date": "05/22/2026", "Quote Date": "05/31/2026"} for _ in range(4)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    for r in out:
        _eff_ge_quote(r)


def test_wh_effective_on_or_after_quote():
    # WH-003
    handler = WhHandler()
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Effective Date": "05/22/2026", "Quote Date": "05/31/2026"} for _ in range(4)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    for r in out:
        _eff_ge_quote(r)


def test_cargo_effective_on_or_after_quote():
    # CARGO-006
    handler = CargoHandler()
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Effective Date": "05/22/2026", "Quote Date": "05/31/2026"} for _ in range(4)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    for r in out:
        _eff_ge_quote(r)


def test_apd_effective_on_or_after_quote():
    # APD-013
    handler = ApdHandler()
    rows = [{"Test ID": "x", "Type of Entity": "Corporation",
             "Effective Date": "05/22/2026", "Quote Date": "05/31/2026"} for _ in range(4)]
    out = handler.post_process(rows, "01_Policy_Info", "")
    for r in out:
        _eff_ge_quote(r)


def test_dw_effective_clamped_to_today_not_before_quote():
    # DEF-023: quote is pinned to today; a past effective date must be pushed up
    # to today so it never precedes the quote.
    handler = DwHandler()
    rows = [{"Test ID": "TS-001", "Product Selected": "Dwelling Fire DP-3",
             "Effective Date": "01/01/2020", "Expiration Date": "",
             "Quote Date": "01/01/2020"}]
    out = handler.post_process(rows, "Policy Info", "")
    today = _dt.date.today()
    assert _parse_date(out[0]["Quote Date"]) == today
    assert _parse_date(out[0]["Effective Date"]) == today
    _eff_ge_quote(out[0])
    # Expiration re-derived from the clamped effective date (today + 1 year).
    assert _parse_date(out[0]["Expiration Date"]) == today.replace(year=today.year + 1)


def test_ho_effective_clamped_to_today_not_before_quote():
    # HO-022
    handler = HoHandler()
    rows = [{"Test ID": "TS-001", "Product Selected": "HO-3 Homeowners",
             "Effective Date": "03/15/2019", "Expiration Date": "",
             "Quote Date": "03/15/2019"}]
    out = handler.post_process(rows, "Policy Info", "")
    today = _dt.date.today()
    assert _parse_date(out[0]["Quote Date"]) == today
    assert _parse_date(out[0]["Effective Date"]) == today
    _eff_ge_quote(out[0])


def test_dw_future_effective_preserved():
    # Guard: a valid future effective date (>= today) is NOT altered.
    handler = DwHandler()
    future = _dt.date.today() + _dt.timedelta(days=45)
    fstr = future.strftime("%m/%d/%Y")
    rows = [{"Test ID": "TS-001", "Product Selected": "Dwelling Fire DP-3",
             "Effective Date": fstr, "Expiration Date": "", "Quote Date": fstr}]
    out = handler.post_process(rows, "Policy Info", "")
    assert _parse_date(out[0]["Effective Date"]) == future


# ---------------------------------------------------------------------------
# Per-LOB open defects (2026-06-30/07-01 tracker)
# ---------------------------------------------------------------------------

def test_im_override_reason_gated_on_any_override():
    # DF-IM-019: Override Reason is mandatory when ANY override field has a value
    # (not just Fee Override), and blank when none do.
    handler = ImHandler()
    rows = [
        # UW Surcharge % present, Fee Override blank -> reason must be populated
        {"Test ID": "x", "UW Surcharge %": "5%", "Fee Override ($)": "",
         "Deductible Override ($)": "", "Override Reason": ""},
        # every override blank -> reason must be cleared
        {"Test ID": "x", "UW Surcharge %": "", "Fee Override ($)": "",
         "Deductible Override ($)": "", "Override Reason": "stray"},
        # a non-fee override present -> reason populated
        {"Test ID": "x", "UW Surcharge %": "", "Fee Override ($)": "",
         "Deductible Override ($)": "250", "Override Reason": ""},
    ]
    out = handler.post_process(rows, "01_Policy_Info", "")
    assert out[0]["Override Reason"].strip()      # surcharge present
    assert out[1]["Override Reason"] == ""         # nothing present
    assert out[2]["Override Reason"].strip()      # deductible override present


def test_dw_loc_number_sequential_per_insured():
    # DEF-021: "Loc #" must run 1..N within each insured on DF Locations.
    handler = DwHandler()
    rows = [{"Test ID": "TS-001", "Loc #": 7, "Street Address": "1 A", "City": "X",
             "State": "VA", "ZIP Code": "00000"}]
    prev = {"Policy Info": [{"Test ID": "TS-001"}]}
    out = handler.post_process(rows, "DF Locations", "", prev)
    locs = [r["Loc #"] for r in out]
    assert locs == list(range(1, len(out) + 1))    # sequential from 1


def test_ho_type_of_entity_varies():
    # HO-021: Type of Entity collapses onto "Individual" for every row.
    handler = HoHandler()
    rows = [{"Test ID": f"TS-00{i}", "Product Selected": "HO-3 Homeowners",
             "Type of Entity": "Individual", "Effective Date": "", "Quote Date": ""}
            for i in range(1, 8)]
    out = handler.post_process(rows, "Policy Info", "")
    seen = {r["Type of Entity"] for r in out}
    assert len(seen) >= 4
    assert seen <= set(_PL_ENTITY_TYPES)


def test_dw_type_of_entity_varies():
    handler = DwHandler()
    rows = [{"Test ID": f"TS-00{i}", "Product Selected": "Dwelling Fire DP-3",
             "Type of Entity": "Individual", "Effective Date": "", "Quote Date": ""}
            for i in range(1, 8)]
    out = handler.post_process(rows, "Policy Info", "")
    assert len({r["Type of Entity"] for r in out}) >= 4


def test_apd_child_counts_vary_across_insureds():
    # APD-014: per-insured schedule counts must not follow a fixed pattern.
    handler = ApdHandler()
    prev = {"01_Policy_Info": [{"Test ID": f"TS-0{i:02d}"} for i in range(1, 9)]}
    rows = [{"Test ID": f"TS-0{i:02d}", "License Number": f"L{i}"} for i in range(1, 9)]
    out = handler.post_process(rows, "03_APD_Drivers", "", prev)
    counts = Counter(r["Test ID"] for r in out)
    assert all(c >= 2 for c in counts.values())     # still multi-row (>= min)
    assert len(set(counts.values())) > 1            # not a single fixed count


def test_cargo_child_counts_one_per_policy():
    # Cargo defect ("for 20 sets generates 60 instead of 20"): each policy carries
    # EXACTLY ONE child record by default, so N test cases yield N records — not a
    # ballooned multi-row schedule. (APD keeps the multi-entry default.)
    handler = CargoHandler()
    prev = {"01_Policy_Info": [{"Test ID": f"TS-0{i:02d}"} for i in range(1, 9)]}
    rows = [{"Test ID": f"TS-0{i:02d}", "Commodity (Select from list)": "Steel"}
            for i in range(1, 9)]
    out = handler.post_process(rows, "05_Cargo_Commodities", "", prev)
    counts = Counter(r["Test ID"] for r in out)
    assert set(counts.values()) == {1}              # exactly one record per policy
    assert len(out) == 8                            # total == number of test cases


def test_apd_child_counts_stay_multi_row():
    # APD keeps the multi-entry default (APD-006..009): several records per insured.
    handler = ApdHandler()
    prev = {"01_Policy_Info": [{"Test ID": f"TS-0{i:02d}"} for i in range(1, 9)]}
    rows = [{"Test ID": f"TS-0{i:02d}", "VIN Number": f"V{i}"} for i in range(1, 9)]
    out = handler.post_process(rows, "04_APD_Vehicles", "", prev)
    counts = Counter(r["Test ID"] for r in out)
    assert all(c >= 3 for c in counts.values())     # multi-row per insured
