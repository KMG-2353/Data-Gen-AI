"""The xlsx download must not crash on control characters an LLM may emit.

Regression: an APD/DW value like 'Water Damage\\x0bNon-Weather Related' (embedded
0x0B) made openpyxl raise 'cannot be used in worksheets' and failed the whole
download. The write boundary now strips illegal control chars.
"""
from io import BytesIO

from openpyxl import load_workbook
from fastapi.testclient import TestClient

from app import main


def test_download_strips_illegal_control_chars():
    client = TestClient(main.app)
    sid = "test-illegal-chars"
    main.sessions[sid] = {
        "headers_by_sheet": {"Policy": ["Test ID", "Type of Loss"]},
        "unique_headers_by_sheet": {"Policy": ["Test ID", "Type of Loss"]},
        "filename": "x.xlsx",
        "policy_type": "GENERIC",
        "generated_data_by_sheet": {
            "Policy": {
                "original_headers": ["Test ID", "Type of Loss"],
                "unique_headers": ["Test ID", "Type of Loss"],
                "data": [
                    {"Test ID": "TS-001",
                     "Type of Loss": "Water Damage\x0bNon-Weather Related"},
                    {"Test ID": "TS-002", "Type of Loss": "Fire"},
                ],
            }
        },
    }
    try:
        resp = client.get(f"/api/download/{sid}")
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Policy"]
        # The illegal char is gone; the surrounding text survives.
        assert ws.cell(row=2, column=2).value == "Water DamageNon-Weather Related"
        assert ws.cell(row=3, column=2).value == "Fire"
    finally:
        main.sessions.pop(sid, None)
