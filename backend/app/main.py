import re

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook, Workbook
from io import BytesIO
import uuid
import os
from dotenv import load_dotenv

from app.llm_service import generate_test_data, detect_policy_type, detect_policy_type_from_headers
from app.policies import get_handler
from app.rulebook import config as rb_config
from app.rulebook.profiles import select_profile
from app.rulebook.scenario import parse_scenarios
from app.rulebook.variety import enforce_variety_fields

# SPG rater family that shares the universal dropdown-variety contract
# (phone/fax format + address-state spread). RRG/IMS/PAP manage their own
# (restricted) states and are intentionally excluded.
_SPG_LOBS = frozenset({"IM", "DW", "HO", "CARGO", "APD"})
from app.policies.ims import (
    canonical_sheet_name as _ims_canonical_sheet_name,
    extract_lob_flags as _ims_extract_lob_flags,
    filter_rows_by_lob as _ims_filter_rows_by_lob,
    enforce_lob_selection as _ims_enforce_lob_selection,
    enforce_state_selection as _ims_enforce_state_selection,
)
from app.policies.pap_quincy import (
    enforce_pap_state_selection as _pap_enforce_state_selection,
    enforce_pap_policy_field_rules as _pap_enforce_field_rules,
)
from app.address_service import generate_verified_addresses as _generate_verified_addresses
from app.llm_service import configure_openai as _configure_openai

load_dotenv()

app = FastAPI()

# Configure CORS with environment-aware origins
_FIXED_ORIGINS = [
    "http://localhost:5173",
    "http://localhost:5174",
    "https://data-gen-ai-1.onrender.com",
]

frontend_url = os.getenv("FRONTEND_URL")
if frontend_url and frontend_url not in _FIXED_ORIGINS:
    _FIXED_ORIGINS.append(frontend_url)

# Allow any device on a local network (192.168.x.x, 10.x.x.x, 172.16-31.x.x)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_FIXED_ORIGINS,
    # Allow localhost/127.0.0.1 on ANY port (Vite picks 5173/5174/5175/… based on
    # what's free) plus any device on a local network.
    allow_origin_regex=r"http://(localhost|127\.0\.0\.1|192\.168\.\d+\.\d+|10\.\d+\.\d+\.\d+|172\.(1[6-9]|2\d|3[01])\.\d+\.\d+):\d+",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Store uploaded file data temporarily
sessions = {}

# @app.get("/api/hello")
# def read_hello():
#     return {"message": "Test123"}

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")

        # Read file content
        content = await file.read()

        # Load workbook from bytes
        workbook = load_workbook(BytesIO(content), read_only=True)

        # Canonicalize sheet names when the upload is an IMS workbook
        # (corrects known misspellings like "LIBIALITY" → "LIABILITY").
        policy_type_for_upload = detect_policy_type(file.filename)

        # Extract headers from first row of all sheets
        headers_by_sheet = {}       # original names (for display & Excel output)
        unique_headers_by_sheet = {} # deduplicated names (for LLM)
        for sheet in workbook.worksheets:
            original_headers = []
            for cell in sheet[1]:
                if cell.value:
                    original_headers.append(str(cell.value))
            if not original_headers:
                continue

            # Make duplicate headers unique by appending column position
            seen = {}
            unique_headers = []
            for idx, h in enumerate(original_headers):
                if original_headers.count(h) > 1:
                    # Track occurrence number
                    seen[h] = seen.get(h, 0) + 1
                    unique_headers.append(f"{h} (Col {idx + 1})")
                else:
                    unique_headers.append(h)

            sheet_title = sheet.title
            if policy_type_for_upload == "IMS":
                sheet_title = _ims_canonical_sheet_name(sheet_title)
            headers_by_sheet[sheet_title] = original_headers
            unique_headers_by_sheet[sheet_title] = unique_headers

        workbook.close()

        if not headers_by_sheet:
            raise HTTPException(status_code=400, detail="No headers found in any sheet")

        # Detect policy type from headers (primary); filename is kept as fallback
        detected_policy_type = detect_policy_type_from_headers(headers_by_sheet)
        if detected_policy_type == "GENERIC":
            detected_policy_type = detect_policy_type(file.filename)

        # Generate session ID
        session_id = str(uuid.uuid4())

        # Store session data
        sessions[session_id] = {
            "headers_by_sheet": headers_by_sheet,
            "unique_headers_by_sheet": unique_headers_by_sheet,
            "filename": file.filename,
            "policy_type": detected_policy_type,
        }

        return {
            "session_id": session_id,
            "headers_by_sheet": headers_by_sheet,
            "sheet_names": list(headers_by_sheet.keys()),
            "sheet_count": len(headers_by_sheet),
            "filename": file.filename
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.post("/api/generate")
async def generate_data(request: dict):
    """Generate test data using LLM for all sheets"""
    session_id = request.get("session_id")
    row_count = request.get("row_count", 10)
    special_instructions = request.get("special_inst", "")
    lob_selection: list[str] = request.get("lob_selection", [])   # e.g. ["Crime", "General Liability"]
    state_selection: list[str] = request.get("state_selection", [])  # e.g. ["CA", "TX"]
    driver_count: int | None = request.get("driver_count")   # PAP: user-selected driver count
    vehicle_count: int | None = request.get("vehicle_count")  # PAP: user-selected vehicle count

    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    unique_headers_by_sheet = sessions[session_id]["unique_headers_by_sheet"]
    original_headers_by_sheet = sessions[session_id]["headers_by_sheet"]
    filename = sessions[session_id].get("filename", "")

    # Use header-based detection stored at upload time.
    # If not stored (session created before code update) re-detect from headers.
    policy_type = sessions[session_id].get("policy_type")
    if not policy_type:
        policy_type = detect_policy_type_from_headers(unique_headers_by_sheet)
        if policy_type == "GENERIC":
            policy_type = detect_policy_type(filename)
        # Cache it so subsequent generate calls for same session are fast
        sessions[session_id]["policy_type"] = policy_type
    handler = get_handler(policy_type)
    print(f"Detected policy_type={policy_type} for filename='{filename}'")

    # -------------------------------------------------------------------------
    # Build augmented_special: user instructions + UI-level LOB/state hints.
    # LOB selection is injected at the TOP so it sits inside USER SPECIAL
    # INSTRUCTIONS in the LLM prompt.  Special instructions from the text box
    # always take precedence because they are written AFTER the LOB hint.
    # -------------------------------------------------------------------------
    augmented_special = special_instructions or ""

    # LOB selection hint — injected BEFORE user text so user text can still override.
    # Format: individual "LOBName = Yes/No" lines so parse_special_instructions
    # regex (lob_name\s*=\s*yes) can correctly populate pi.active_lobs.
    if lob_selection and policy_type == "IMS":
        all_ims_lobs = [
            "General Liability", "Property", "Crime",
            "Inland Marine", "Commercial Auto", "Optional Coverage",
        ]
        selected_lower = {s.lower().strip() for s in lob_selection}
        lob_lines = []
        for lob in all_ims_lobs:
            val = "Yes" if lob.lower() in selected_lower else "No"
            lob_lines.append(f"{lob} = {val}")
        lob_hint = "SELECTED LOBs (Policy sheet): " + ", ".join(lob_lines) + "."
        augmented_special = f"{lob_hint}\n{augmented_special}".strip()

    # State selection hint — format matches the parse_special_instructions regex:
    # (?:only|for|states?)[:\s]+([A-Z]{2}...)  so pi.allowed_states gets populated
    # and the STRUCTURED CONSTRAINTS block outputs "STATES: Only generate data for: ..."
    if state_selection:
        states_str = ", ".join(state_selection)
        si_lower = augmented_special.lower()
        if not any(s.lower() in si_lower for s in state_selection):
            state_hint = f"states: {states_str}."
            augmented_special = f"{augmented_special}\n{state_hint}".strip() if augmented_special else state_hint

    # PAP: inject user-selected driver/vehicle combination hint.
    # This tells the LLM exactly which combo to use in Test Case Details.
    # Goes at the TOP so it takes highest precedence over any default behavior.
    if driver_count and vehicle_count and policy_type == "PAP":
        combo_hint = (
            f"DRIVER/VEHICLE COMBINATION: Use exactly {driver_count} Driver & {vehicle_count} Vehicle "
            f"for ALL test cases. Test Case Details must be \"{driver_count} Driver & {vehicle_count} Vehicle\"."
        )
        augmented_special = f"{combo_hint}\n{augmented_special}".strip() if augmented_special else combo_hint

    # Rulebook engine: select the profile (via existing detection) and parse any
    # scenarios from the special instructions. The handlers self-parse the same
    # text at their count/validate seams; here we only surface the profile and
    # any cap adjustments (R11) back to the caller. The whole path is gated by
    # the RULEBOOK_ENABLED fallback flag.
    scenario_adjustments: list[str] = []
    scenario_insured_count = 0
    if rb_config.RULEBOOK_ENABLED:
        _scenarios = parse_scenarios(augmented_special)
        scenario_adjustments = list(_scenarios.adjustments)
        scenario_insured_count = len(_scenarios.specs)
        if _scenarios.specs:
            print(
                f"[rulebook] profile={select_profile(policy_type).policy_type} "
                f"scenario_insureds={scenario_insured_count} "
                f"adjustments={scenario_adjustments}"
            )

    try:
        # PAP: pre-generate Census-verified addresses before any sheet LLM call.
        # The post-processor will overwrite all address fields from this map —
        # the main LLM calls play no role in address generation.
        _pap_verified_addresses: dict[str, dict] = {}
        if policy_type == "PAP" and state_selection:
            try:
                _provider = os.getenv("MODEL_PROVIDER", "openai").lower()
                _model = os.getenv(
                    "GEMINI_MODEL" if _provider == "gemini" else "OPENAI_MODEL",
                    "gpt-4o",
                )
                _pap_verified_addresses = _generate_verified_addresses(
                    state_selection,
                    n_groups=row_count,
                    client=_configure_openai(),
                    model=_model,
                )
            except Exception as _addr_exc:
                print(f"[PAP] Address pre-generation failed (will fall back to LLM values): {_addr_exc}")

        # Generate data for ALL sheets sequentially, passing previous data for consistency
        generated_data_by_sheet = {}
        previous_sheets_data = {}

        # Track policy, driver, and vehicle data for cross-sheet row expansion
        policy_data = None
        driver_data = None
        vehicle_data = None

        # IMS-only: LOB flags extracted from the Policy sheet so we can filter
        # sub-sheet rows where the LOB is toggled off.
        lob_flags: dict[int, dict[str, bool]] = {}

        # Process sheets in dependency order regardless of workbook sheet order:
        #   policy → driver/vehicle/infraction → summary/policy_info/questions → assignment
        # Summary, Policy Info, and Questions sheets all need policy_data, so they
        # must come after Policy. Assignment needs driver+vehicle, so it stays last.
        _SHEET_ORDER = {
            "policy": 0, "driver": 1, "vehicle": 2, "infraction": 3,
            "summary": 4, "policy_info": 5, "questions_remarks": 6, "assignment": 7,
        }
        sheet_items = sorted(
            list(unique_headers_by_sheet.items()),
            key=lambda x: _SHEET_ORDER.get(handler.detect_sheet_type(x[0]), 3),
        )

        for sheet_name, unique_headers in sheet_items:
            original_headers = original_headers_by_sheet[sheet_name]

            # Handler decides row count and sheet-specific prompt augmentation
            sheet_type = handler.detect_sheet_type(sheet_name)
            effective_row_count, additional_rules = handler.build_sheet_context(
                sheet_name=sheet_name,
                policy_data=policy_data,
                driver_data=driver_data,
                original_row_count=row_count,
                special_instruction=augmented_special,
            )

            print(f"Generating {effective_row_count} rows for sheet '{sheet_name}' (type={sheet_type}) with headers: {unique_headers}")

            # Deterministic pre-generation path (e.g. PAP assignment sheet).
            pre_generated = handler.pre_generate(
                sheet_name=sheet_name,
                unique_headers=unique_headers,
                policy_data=policy_data,
                driver_data=driver_data,
                vehicle_data=vehicle_data,
                previous_sheets_data=previous_sheets_data if previous_sheets_data else None,
            )

            if pre_generated is not None:
                print(f"Using deterministic pre-generation for sheet '{sheet_name}'")
                data = pre_generated
            elif effective_row_count == 0:
                # Zero-row sheets (e.g. infraction with no drivers flagged)
                data = []
            else:
                data = generate_test_data(
                    headers=unique_headers,
                    row_count=row_count,
                    special_instruction=augmented_special,
                    sheet_name=sheet_name,
                    previous_sheets_data=previous_sheets_data if previous_sheets_data else None,
                    row_count_override=effective_row_count,
                    additional_rules=additional_rules,
                )

            # Run handler-specific post-processing (IMS sheet enforcers +
            # cross-sheet consistency; generic/PAP are no-ops here).
            data = handler.post_process(
                rows=data,
                sheet_name=sheet_name,
                special_instruction=augmented_special,
                previous_sheets_data=previous_sheets_data if previous_sheets_data else None,
            )

            # SPG family (IM/DW/HO/Cargo/APD): universal dropdown-variety pass —
            # snap phone/fax to U.S. format and fan address-state columns across
            # the full US-state pool (never collapsing onto the binding state).
            # Runs AFTER the handler so dependency-blanked cells stay blank;
            # honours a UI state filter when present. Deterministic summary sheets
            # are skipped (their state mirrors the binding state by design).
            if (
                rb_config.RULEBOOK_ENABLED
                and policy_type in _SPG_LOBS
                and sheet_type != "scenario_details"
                and data
            ):
                data = enforce_variety_fields(data, state_selection=state_selection)

            # IMS: once the Policy sheet is done, extract LOB flags so later
            # sub-sheets can be filtered (Defect #216: LOB=No → drop row).
            if policy_type == "IMS" and sheet_type == "policy":
                # Enforce UI LOB selection before extracting flags (special instructions win)
                if lob_selection:
                    data = _ims_enforce_lob_selection(data, lob_selection, special_instructions)
                    print(f"LOB selection enforced on '{sheet_name}': {lob_selection}")
                lob_flags = _ims_extract_lob_flags(data)
                print(f"Extracted LOB flags from '{sheet_name}': {len(lob_flags)} rows")

            # IMS: filter sub-sheet rows against the LOB toggles.
            if policy_type == "IMS" and lob_flags:
                original_count = len(data)
                data = _ims_filter_rows_by_lob(data, sheet_name, lob_flags)
                if len(data) < original_count:
                    print(
                        f"LOB filtering: removed {original_count - len(data)} "
                        f"rows from '{sheet_name}' (LOB=No)"
                    )

            # IMS: deterministically enforce state selection across all sheets.
            # This is a hard post-processing step — LLM prompt alone is unreliable.
            if policy_type == "IMS" and state_selection and data:
                data = _ims_enforce_state_selection(data, sheet_name, state_selection)
                print(f"State selection enforced on '{sheet_name}': {state_selection}")

            # PAP: deterministically enforce state selection + verified addresses.
            if policy_type == "PAP" and state_selection and data:
                data = _pap_enforce_state_selection(
                    data, sheet_name, state_selection,
                    policy_rows=policy_data,
                    pre_generated_addresses=_pap_verified_addresses,
                )
                print(f"PAP state selection enforced on '{sheet_name}': {state_selection}")

            # PAP: enforce field-level rules (UMPD blank for ME, endorsement >= effective).
            if policy_type == "PAP" and sheet_type == "policy" and data:
                data = _pap_enforce_field_rules(data)
                print(f"PAP field rules enforced on '{sheet_name}'")

            generated_data_by_sheet[sheet_name] = {
                "original_headers": original_headers,
                "unique_headers": unique_headers,
                "data": data
            }

            # Add this sheet's data to context for subsequent sheets
            previous_sheets_data[sheet_name] = data

            # Track policy, driver, and vehicle data for downstream sheets
            if sheet_type == "policy":
                policy_data = data
            elif sheet_type == "driver":
                driver_data = data
            elif sheet_type == "vehicle":
                vehicle_data = data

            print(f"Generated {len(data)} rows for sheet '{sheet_name}' successfully")

        # Restore original workbook sheet order for the Excel output
        generated_data_by_sheet = {
            name: generated_data_by_sheet[name]
            for name in unique_headers_by_sheet
            if name in generated_data_by_sheet
        }

        # Store generated data in session
        sessions[session_id]["generated_data_by_sheet"] = generated_data_by_sheet

        return {
            "session_id": session_id,
            "sheets_generated": list(generated_data_by_sheet.keys()),
            "row_count_per_sheet": row_count,
            "scenario_insured_count": scenario_insured_count,
            "scenario_adjustments": scenario_adjustments,
            "status": "complete"
        }
    except Exception as e:
        import traceback
        print(f"Error generating data: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error generating data: {str(e)}")


@app.get("/api/download/{session_id}")
async def download_excel(session_id: str):
    """Download generated data as Excel file with all sheets"""
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]

    if "generated_data_by_sheet" not in session:
        raise HTTPException(status_code=400, detail="No generated data found")

    generated_data_by_sheet = session["generated_data_by_sheet"]

    # Create Excel workbook
    workbook = Workbook()
    # Remove default sheet
    workbook.remove(workbook.active)

    # Create a sheet for each generated dataset
    for sheet_name, sheet_data in generated_data_by_sheet.items():
        sheet = workbook.create_sheet(title=sheet_name[:31])  # Excel limits sheet names to 31 chars
        original_headers = sheet_data["original_headers"]
        unique_headers = sheet_data["unique_headers"]
        data = sheet_data["data"]

        # Write original headers in row 1
        for col, header in enumerate(original_headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # Write data rows using unique header keys to look up values
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, unique_header in enumerate(unique_headers, 1):
                value = row_data.get(unique_header, "")
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"test_data_{session_id[:8]}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
